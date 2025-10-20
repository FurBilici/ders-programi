# app.py
import streamlit as st
import pandas as pd
import json, os, io, textwrap
from io import BytesIO
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

st.set_page_config(page_title="Ders Programı", layout="wide")
STATE_FILE = "timetable_state.json"

# ====================== Yardımcılar ======================

def default_constraint_settings():
    return {
        "online_cap": 3,
        "max_per_room": 1,
        "enf_instructor_no_overlap": True,
        "enf_class_no_overlap": True,
    }

def ensure_session_defaults():
    if "days" not in st.session_state:
        st.session_state.days = ["Pzt","Sal","Çar","Per","Cum"]
    if "slots_per_day" not in st.session_state:
        st.session_state.slots_per_day = 10
    if "time_labels" not in st.session_state:
        st.session_state.time_labels = {
            0:"08:45",1:"09:35",2:"10:25",3:"11:15",4:"12:05",
            5:"13:40",6:"14:30",7:"15:20",8:"16:10",9:"17:00"
        }
    if "rooms" not in st.session_state:
        st.session_state.rooms = [{"id":"Oda-1"},{"id":"Oda-2"}]
    if "instructors" not in st.session_state:
        st.session_state.instructors = ["Hoca_A","Hoca_B"]
    if "instructor_unavailable" not in st.session_state:
        st.session_state.instructor_unavailable = {h: set() for h in st.session_state.instructors}
    if "courses" not in st.session_state:
        st.session_state.courses = [
            {"id":"SAN1101","ad":"Eski Anadolu Uygarlıkları I","hoca":"Hoca_A","sinif":1,"sure":3,"ardisik":True,"online":False},
            {"id":"SAN1102","ad":"Sanat Tarihine Giriş","hoca":"Hoca_B","sinif":1,"sure":3,"ardisik":True,"online":False},
            {"id":"KAR100","ad":"Kariyer Planlama","hoca":"Hoca_A","sinif":1,"sure":1,"ardisik":False,"online":True},
        ]
    if "constraint_settings" not in st.session_state:
        st.session_state.constraint_settings = default_constraint_settings()
    if "day_start_slot" not in st.session_state:
        st.session_state.day_start_slot = {i: 0 for i in range(len(st.session_state.days))}
    if "day_use_slots" not in st.session_state:
        st.session_state.day_use_slots = {i: st.session_state.slots_per_day for i in range(len(st.session_state.days))}
    if "pins" not in st.session_state:
        st.session_state.pins = []
    if "strategy" not in st.session_state:
        st.session_state.strategy = "Kıtlık-önce (önerilir)"

def _to_bool(v):
    if isinstance(v, bool): return v
    if v is None: return False
    s = str(v).strip().lower()
    return s in ["true","1","evet","yes","y","t","e","doğru","on"]

def save_state():
    data = {
        "days": st.session_state.days,
        "slots_per_day": st.session_state.slots_per_day,
        "time_labels": st.session_state.time_labels,
        "rooms": st.session_state.rooms,
        "instructors": st.session_state.instructors,
        "instructor_unavailable": {k: list(v) for k, v in st.session_state.instructor_unavailable.items()},
        "courses": st.session_state.courses,
        "constraint_settings": st.session_state.constraint_settings,
        "day_start_slot": st.session_state.day_start_slot,
        "day_use_slots": st.session_state.day_use_slots,
        "pins": st.session_state.pins,
        "strategy": st.session_state.strategy,
    }
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_state():
    if not os.path.exists(STATE_FILE): return False
    with open(STATE_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    st.session_state.days = data["days"]
    st.session_state.slots_per_day = int(data["slots_per_day"])
    st.session_state.time_labels = {int(k): v for k, v in data["time_labels"].items()}
    st.session_state.rooms = data["rooms"]
    st.session_state.instructors = data["instructors"]
    st.session_state.instructor_unavailable = {k: set(map(tuple, v)) for k, v in data["instructor_unavailable"].items()}
    st.session_state.courses = data["courses"]
    st.session_state.constraint_settings = data.get("constraint_settings", default_constraint_settings())
    dcount = len(st.session_state.days)
    st.session_state.day_start_slot = {int(k): int(v) for k, v in data.get("day_start_slot", {i:0 for i in range(dcount)}).items()}
    st.session_state.day_use_slots  = {int(k): int(v) for k, v in data.get("day_use_slots",  {i:st.session_state.slots_per_day for i in range(dcount)}).items()}
    st.session_state.pins = data.get("pins", [])
    st.session_state.strategy = data.get("strategy", "Kıtlık-önce (önerilir)")
    return True

def export_courses_csv(courses):
    out = io.StringIO()
    cols = ["id","ad","hoca","sinif","sure","ardisik","online"]
    pd.DataFrame([{k:c.get(k,"") for k in cols} for c in courses], columns=cols).to_csv(out, index=False)
    return out.getvalue()

def export_courses_xlsx(courses):
    cols = ["id","ad","hoca","sinif","sure","ardisik","online"]
    df = pd.DataFrame([{k:c.get(k,"") for k in cols} for c in courses], columns=cols)
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="dersler", index=False)
    bio.seek(0)
    return bio

# ====================== Kıtlık Hesabı ======================

def count_feasible_starts_for_course(c, day_start_slot, day_use_slots, spd, inst_unav, days_len):
    L = int(c["sure"])
    h = c["hoca"]
    if L <= 0: return 0
    feas = 0
    for d in range(days_len):
        start0 = int(day_start_slot.get(d, 0))
        use0   = int(day_use_slots.get(d, spd))
        end_allowed = min(spd, start0 + use0) - 1
        if end_allowed < start0 or L > (end_allowed - start0 + 1):
            continue
        for s in range(start0, end_allowed - L + 2):
            if any((d, ss) in inst_unav.get(h, set()) for ss in range(s, s+L)):
                continue
            feas += 1
    return feas

# ====================== Greedy Planlayıcı (Gün-Gün) + PIN ======================

def greedy_schedule(days, spd, rooms, courses, inst_unav, cs, day_start_slot, day_use_slots, pins, strategy):
    n_days = len(days)
    n_rooms = len(rooms)
    online_cap = int(cs["online_cap"])
    max_per_room = int(cs["max_per_room"])
    enf_inst = bool(cs["enf_instructor_no_overlap"])
    enf_class = bool(cs["enf_class_no_overlap"])

    room_occ = [[[0]*spd for _ in range(n_rooms)] for __ in range(n_days)]
    online_load = [[0]*spd for _ in range(n_days)]
    busy_inst = [[set() for _ in range(spd)] for __ in range(n_days)]
    busy_class = [[set() for _ in range(spd)] for __ in range(n_days)]

    placed, unplaced = [], []

    idx_by_id = {c["id"]: i for i, c in enumerate(courses)}

    # ---- 1) PIN'ler ----
    pinned_ci = set()
    for p in pins:
        cid = p.get("id", "").strip()
        if cid not in idx_by_id:  # bilinmeyen ders
            continue
        ci = idx_by_id[cid]
        if ci in pinned_ci:
            continue
        c = courses[ci]; L = int(c["sure"])
        d = int(p.get("day", 0))
        start = int(p.get("start", 0))
        channel = p.get("channel", "FaceToFace")
        room_id = p.get("room", None)

        start0 = int(day_start_slot.get(d, 0))
        use0   = int(day_use_slots.get(d, spd))
        end_allowed = min(spd, start0 + use0) - 1
        if start < start0 or start + L - 1 > end_allowed or start + L - 1 >= spd:
            unplaced.append((ci, f"PIN geçersiz: gün penceresi dışında ({days[d]} {start0}-{end_allowed})"))
            continue

        if any((d, s) in inst_unav.get(c["hoca"], set()) for s in range(start, start+L)):
            unplaced.append((ci, "PIN geçersiz: hoca uygunsuz saat"))
            continue

        if enf_inst and any((c["hoca"] in busy_inst[d][s]) for s in range(start, start+L)):
            unplaced.append((ci, "PIN geçersiz: hoca çakışması"))
            continue
        if enf_class and any((c["sinif"] in busy_class[d][s]) for s in range(start, start+L)):
            unplaced.append((ci, "PIN geçersiz: sınıf çakışması"))
            continue

        if channel == "Online" or c["online"]:
            if any(online_load[d][s] >= online_cap for s in range(start, start+L)):
                unplaced.append((ci, "PIN geçersiz: online kapasite dolu"))
                continue
            for s in range(start, start+L):
                online_load[d][s] += 1
                busy_inst[d][s].add(c["hoca"])
                busy_class[d][s].add(c["sinif"])
            placed.append((ci, d, start, "Online", "ONLINE"))
            pinned_ci.add(ci)
        else:
            if not room_id:
                unplaced.append((ci, "PIN geçersiz: oda belirtilmemiş"))
                continue
            try:
                ri = [r["id"] for r in rooms].index(room_id)
            except ValueError:
                unplaced.append((ci, f"PIN geçersiz: oda bulunamadı ({room_id})"))
                continue
            if any(room_occ[d][ri][s] >= max_per_room for s in range(start, start+L)):
                unplaced.append((ci, "PIN geçersiz: oda kapasitesi dolu"))
                continue
            for s in range(start, start+L):
                room_occ[d][ri][s] += 1
                busy_inst[d][s].add(c["hoca"])
                busy_class[d][s].add(c["sinif"])
            placed.append((ci, d, start, "FaceToFace", rooms[ri]["id"]))
            pinned_ci.add(ci)

    # ---- 2) Sıralama ----
    idx_offline = [i for i,c in enumerate(courses) if (not c["online"]) and i not in pinned_ci]
    idx_online  = [i for i,c in enumerate(courses) if c["online"] and i not in pinned_ci]

    def scarcity_key(i):
        c = courses[i]
        feas = count_feasible_starts_for_course(c, day_start_slot, day_use_slots, spd, inst_unav, n_days)
        return (feas, -int(c.get("sinif", 1) == 4), -int(c["sure"]))

    if strategy.startswith("Kıtlık"):
        idx_offline.sort(key=scarcity_key)
        idx_online.sort(key=scarcity_key)
    else:
        idx_offline.sort(key=lambda i: -int(courses[i]["sure"]))
        idx_online.sort(key=lambda i: -int(courses[i]["sure"]))

    # ---- 3) Yerleştirme: OFFLINE ----
    for ci in idx_offline:
        c = courses[ci]; L = int(c["sure"])
        done = False
        for d in range(n_days):
            start0 = int(day_start_slot.get(d, 0))
            use0   = int(day_use_slots.get(d, spd))
            end_allowed = min(spd, start0 + use0) - 1
            if end_allowed < start0 or L > (end_allowed - start0 + 1):
                continue
            for start in range(start0, end_allowed - L + 2):
                if any((d, s) in inst_unav.get(c["hoca"], set()) for s in range(start, start+L)):
                    continue
                if enf_inst and any((c["hoca"] in busy_inst[d][s]) for s in range(start, start+L)):
                    continue
                if enf_class and any((c["sinif"] in busy_class[d][s]) for s in range(start, start+L)):
                    continue
                chosen_ri = None
                for ri in range(n_rooms):
                    if any(room_occ[d][ri][s] >= max_per_room for s in range(start, start+L)):
                        continue
                    chosen_ri = ri; break
                if chosen_ri is None:
                    continue
                for s in range(start, start+L):
                    room_occ[d][chosen_ri][s] += 1
                    busy_inst[d][s].add(c["hoca"])
                    busy_class[d][s].add(c["sinif"])
                placed.append((ci, d, start, "FaceToFace", rooms[chosen_ri]["id"]))
                done = True
                break
            if done: break
        if not done:
            unplaced.append((ci, "Uygun oda/slot (gün penceresi içinde) bulunamadı"))

    # ---- 4) Yerleştirme: ONLINE ----
    for ci in idx_online:
        c = courses[ci]; L = int(c["sure"])
        done = False
        for d in range(n_days):
            start0 = int(day_start_slot.get(d, 0))
            use0   = int(day_use_slots.get(d, spd))
            end_allowed = min(spd, start0 + use0) - 1
            if end_allowed < start0 or L > (end_allowed - start0 + 1):
                continue
            for start in range(start0, end_allowed - L + 2):
                if any((d, s) in inst_unav.get(c["hoca"], set()) for s in range(start, start+L)):
                    continue
                if enf_inst and any((c["hoca"] in busy_inst[d][s]) for s in range(start, start+L)):
                    continue
                if enf_class and any((c["sinif"] in busy_class[d][s]) for s in range(start, start+L)):
                    continue
                if any(online_load[d][s] >= int(online_cap) for s in range(start, start+L)):
                    continue
                for s in range(start, start+L):
                    online_load[d][s] += 1
                    busy_inst[d][s].add(c["hoca"])
                    busy_class[d][s].add(c["sinif"])
                placed.append((ci, d, start, "Online", "ONLINE"))
                done = True
                break
            if done: break
        if not done:
            unplaced.append((ci, "Online kapasite/çakışma (gün penceresi)"))

    # ---- 5) Görsel tablo verisi ----
    placed_by_cell = defaultdict(list)
    for ci, d, start, ch, rm in placed:
        L = int(courses[ci]["sure"])
        for s in range(start, start+L):
            placed_by_cell[(d, s, ch, rm)].append(ci)

    rows = []
    for d in range(n_days):
        for s in range(st.session_state.slots_per_day):
            for ri in range(n_rooms):
                key = (d, s, "FaceToFace", rooms[ri]["id"])
                cis = placed_by_cell.get(key, [])
                cell_txt = "-"
                if cis:
                    parts = []
                    for ci in cis:
                        c = courses[ci]
                        parts.append(f"{c['id']} | {c['ad']} | {c['hoca']} | S{c['sinif']}")
                    cell_txt = " / ".join(parts)
                rows.append([st.session_state.days[d], st.session_state.time_labels.get(s, str(s+1)),
                             "FaceToFace", rooms[ri]["id"], cell_txt])
            key = (d, s, "Online", "ONLINE")
            cis = placed_by_cell.get(key, [])
            cell_txt = "-"
            if cis:
                parts = []
                for ci in cis:
                    c = courses[ci]
                    parts.append(f"{c['id']} | {c['ad']} | {c['hoca']} | S{c['sinif']}")
                cell_txt = " / ".join(parts)
            rows.append([st.session_state.days[d], st.session_state.time_labels.get(s, str(s+1)),
                         "Online", "ONLINE", cell_txt])

    timetable_df = pd.DataFrame(rows, columns=["Day","Slot","Channel","Room","Courses"])

    diag_rows = []
    for ci, reason in unplaced:
        c = courses[ci]
        diag_rows.append({
            "id": c["id"], "ad": c["ad"], "hoca": c["hoca"], "sinif": c["sinif"],
            "sure": c["sure"], "online": c["online"], "neden": reason
        })
    diag_df = pd.DataFrame(diag_rows, columns=["id","ad","hoca","sinif","sure","online","neden"])

    return timetable_df, diag_df, placed, unplaced

# ====================== Gün Gün Okunur Tablo ======================

def render_day_tables(timetable_df, days, rooms, time_labels):
    def fmt(cell):
        if not isinstance(cell, str) or cell.strip() == "-" or cell.strip() == "":
            return ""
        parts = [p.strip() for p in cell.split("/") if p.strip()]
        return "\n".join(parts)
    max_slot_index = max(time_labels.keys()) if time_labels else 0
    for d in days:
        cols = ["Saat"] + [r["id"] for r in rooms] + ["ONLINE"]
        rows = []
        for s in range(max_slot_index + 1):
            saat = time_labels.get(s, f"{s+1}. Slot")
            row = [saat]
            for r in rooms:
                mask = (
                    (timetable_df["Day"] == d) &
                    (timetable_df["Slot"] == saat) &
                    (timetable_df["Channel"] == "FaceToFace") &
                    (timetable_df["Room"] == r["id"])
                )
                vals = timetable_df.loc[mask, "Courses"].values
                row.append(fmt(vals[0]) if len(vals) else "")
            mask_on = (
                (timetable_df["Day"] == d) &
                (timetable_df["Slot"] == saat) &
                (timetable_df["Channel"] == "Online") &
                (timetable_df["Room"] == "ONLINE")
            )
            vals_on = timetable_df.loc[mask_on, "Courses"].values
            row.append(fmt(vals_on[0]) if len(vals_on) else "")
            rows.append(row)
        day_df = pd.DataFrame(rows, columns=cols)
        st.markdown(f"### {d}")
        st.table(day_df.style.set_properties(**{"white-space": "pre-wrap"}))

# ====================== PDF Üretimi (wrap + dinamik satır) ======================

def _wrap_cell(text, max_chars):
    if text is None: return ""
    t = str(text).strip()
    if t == "-" or t == "": return ""
    # Var olan "/" ayraçlarını satır kır, sonra her satırı sar
    t = t.replace(" / ", "\n")
    lines = []
    for part in t.split("\n"):
        # boşsa koru
        if not part.strip():
            lines.append("")
            continue
        wrapped = textwrap.wrap(part, width=max_chars, break_long_words=True, break_on_hyphens=True)
        lines.extend(wrapped if wrapped else [part])
    return "\n".join(lines)

def timetable_to_pdf(timetable_df, days, rooms, time_labels, pdf_path):
    max_slot_index = max(time_labels.keys()) if time_labels else 0
    # Sütun genişlik oranları: Saat dar, diğerleri eşit
    n_content = len(rooms) + 1  # rooms + ONLINE
    # toplam 1.0: Saat=0.12, kalan eşit pay
    saat_w = 0.12
    rest_w = (1.0 - saat_w) / n_content
    col_widths = [saat_w] + [rest_w]*(n_content)

    # Kolona göre karakter limiti (yaklaşık hesap): genişlik*100 karakter
    # (dpi/font etkileri nedeniyle yaklaşık, ama pratikte iyi sonuç verir)
    col_char_limits = [8] + [max(16, int(rest_w*100))]*n_content

    with PdfPages(pdf_path) as pdf:
        for d in days:
            # Gün DataFrame'i
            cols = ["Saat"] + [r["id"] for r in rooms] + ["ONLINE"]
            rows = []
            for s in range(max_slot_index + 1):
                saat = time_labels.get(s, f"{s+1}. Slot")
                row = [saat]
                for r in rooms:
                    mask = (
                        (timetable_df["Day"] == d) &
                        (timetable_df["Slot"] == saat) &
                        (timetable_df["Channel"] == "FaceToFace") &
                        (timetable_df["Room"] == r["id"])
                    )
                    vals = timetable_df.loc[mask, "Courses"].values
                    val = "" if len(vals)==0 else vals[0]
                    row.append(val)
                mask_on = (
                    (timetable_df["Day"] == d) &
                    (timetable_df["Slot"] == saat) &
                    (timetable_df["Channel"] == "Online") &
                    (timetable_df["Room"] == "ONLINE")
                )
                vals_on = timetable_df.loc[mask_on, "Courses"].values
                val_on = "" if len(vals_on)==0 else vals_on[0]
                row.append(val_on)
                rows.append(row)

            df = pd.DataFrame(rows, columns=cols)

            # Wrap uygulanmış hücre matrisini hazırla ve satır başına max satır sayısını ölç
            wrapped = []
            row_max_lines = []
            for r_i in range(len(df)):
                wr_row = []
                max_lines = 1
                for c_i, col in enumerate(df.columns):
                    raw = df.iat[r_i, c_i]
                    w = _wrap_cell(raw, col_char_limits[c_i])
                    wr_row.append(w)
                    max_lines = max(max_lines, w.count("\n")+1 if w else 1)
                wrapped.append(wr_row)
                row_max_lines.append(max_lines)

            # Şekil
            fig = plt.figure(figsize=(11.69, 8.27))  # A4 yatay
            ax = plt.gca()
            ax.axis('off')
            ax.set_title(f"{d} - Ders Programı", pad=12)

            tbl = ax.table(cellText=wrapped, colLabels=df.columns, loc='center', cellLoc='left')
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(7)  # daha küçük font

            # Sütun genişliklerini uygula (tüm hücreler)
            for (r, c), cell in tbl.get_celld().items():
                # header = r == 0
                w = col_widths[c] if c < len(col_widths) else rest_w
                cell.set_width(w)

            # Satır yüksekliklerini max satır sayısına göre ayarla
            base_h = 0.04  # temel yükseklik
            for r in range(1, len(df)+1):  # 1..n (0 header)
                lines = row_max_lines[r-1]
                h = base_h * max(1.0, lines*0.9)
                for c in range(len(cols)):
                    tbl[(r, c)].set_height(h)

            # Header stil
            for c in range(len(cols)):
                hdr = tbl[(0, c)]
                hdr.set_height(0.05)
                hdr.set_fontsize(8)

            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

# ====================== Excel Üretimi (gün başına ayrı sayfa) ======================

def timetable_to_excel_bytes(timetable_df, days, rooms, time_labels):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    # Varsayılan ilk sayfayı kaldır
    wb.remove(wb.active)

    max_slot_index = max(time_labels.keys()) if time_labels else 0
    room_ids = [r["id"] for r in rooms]

    for d in days:
        ws = wb.create_sheet(title=d)
        headers = ["Saat"] + room_ids + ["ONLINE"]
        ws.append(headers)
        # Bold header
        for col in range(1, len(headers)+1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        for s in range(max_slot_index + 1):
            saat = time_labels.get(s, f"{s+1}. Slot")
            row_vals = [saat]
            # Odalar
            for rid in room_ids:
                mask = (
                    (timetable_df["Day"] == d) &
                    (timetable_df["Slot"] == saat) &
                    (timetable_df["Channel"] == "FaceToFace") &
                    (timetable_df["Room"] == rid)
                )
                vals = timetable_df.loc[mask, "Courses"].values
                v = "" if len(vals)==0 or str(vals[0]).strip()=="-" else str(vals[0]).replace(" / ", "\n")
                row_vals.append(v)
            # ONLINE
            mask_on = (
                (timetable_df["Day"] == d) &
                (timetable_df["Slot"] == saat) &
                (timetable_df["Channel"] == "Online") &
                (timetable_df["Room"] == "ONLINE")
            )
            vals_on = timetable_df.loc[mask_on, "Courses"].values
            v_on = "" if len(vals_on)==0 or str(vals_on[0]).strip()=="-" else str(vals_on[0]).replace(" / ", "\n")
            row_vals.append(v_on)
            ws.append(row_vals)

        # Sar/ortala ve kolon genişlikleri
        wrap = Alignment(wrap_text=True, vertical="top")
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                cell.alignment = wrap

        # Kolon genişlikleri (yaklaşık): Saat dar, diğerleri geniş
        ws.column_dimensions["A"].width = 10
        for idx in range(2, len(headers)+1):
            ws.column_dimensions[chr(64+idx)].width = 45  # B.. son

        # Satır yüksekliğini biraz artır (dolguya göre Excel kendisi de büyütür)
        for rr in range(2, ws.max_row+1):
            ws.row_dimensions[rr].height = 30

    bio = BytesIO()
    from openpyxl.writer.excel import save_workbook
    save_workbook(wb, bio)
    bio.seek(0)
    return bio

# ====================== Uygulama UI ======================

ensure_session_defaults()

left, right = st.columns([0.48, 0.52])

with left:
    st.header("Veri Girişi")

    colA, colB = st.columns(2)
    with colA:
        if st.button("💾 Kaydet (JSON)"):
            save_state()
            st.success("Durum kaydedildi.")
    with colB:
        if st.button("📂 Yükle (JSON)"):
            if load_state():
                st.success("Kayıtlı durum yüklendi.")
                st.rerun()
            else:
                st.warning("Kayıt dosyası bulunamadı.")

    with st.expander("📥 Dersleri İçe/Dışa Aktar", expanded=False):
        template_cols = ["id","ad","hoca","sinif","sure","ardisik","online"]
        template_df = pd.DataFrame([{
            "id":"SAN1101","ad":"Eski Anadolu Uygarlıkları I","hoca":"Hoca_A",
            "sinif":1,"sure":3,"ardisik":True,"online":False
        }], columns=template_cols)
        t_csv = template_df.to_csv(index=False)
        st.download_button("📄 Şablon (CSV) indir", data=t_csv, file_name="ders_sablon.csv", mime="text/csv")
        t_xlsx = export_courses_xlsx(template_df.to_dict(orient="records"))
        st.download_button("📊 Şablon (Excel) indir", data=t_xlsx.getvalue(), file_name="ders_sablon.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.markdown("---")
        col_e1, col_e2 = st.columns(2)
        with col_e1:
            st.download_button("Mevcut dersleri **CSV** indir",
                               data=export_courses_csv(st.session_state.courses),
                               file_name="dersler.csv", mime="text/csv")
        with col_e2:
            st.download_button("Mevcut dersleri **Excel** indir",
                               data=export_courses_xlsx(st.session_state.courses).getvalue(),
                               file_name="dersler.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        uploaded = st.file_uploader("Excel (.xlsx) veya CSV yükle", type=["xlsx","csv"])
        replace_all = st.checkbox("Mevcut listeyi SİL (tam yerine yaz)", value=False)
        update_existing = st.checkbox("Aynı ID'li dersi güncelle", value=True)
        if st.button("İçe Aktar"):
            if not uploaded:
                st.warning("Önce bir dosya yükleyin.")
            else:
                try:
                    if uploaded.name.lower().endswith(".csv"):
                        df = pd.read_csv(uploaded)
                    else:
                        df = pd.read_excel(uploaded, sheet_name=0)
                    df.columns = [c.lower() for c in df.columns]
                    required = set(template_cols)
                    missing = required - set(df.columns)
                    if missing:
                        st.error(f"Eksik kolon(lar): {', '.join(sorted(missing))}")
                    else:
                        df["sinif"] = df["sinif"].astype(int)
                        df["sure"]  = df["sure"].astype(int)
                        df["ardisik"] = df["ardisik"].apply(_to_bool)
                        df["online"]  = df["online"].apply(_to_bool)
                        for h in sorted(set(df["hoca"].dropna())):
                            if h not in st.session_state.instructors:
                                st.session_state.instructors.append(h)
                                st.session_state.instructor_unavailable[h] = set()
                        new_courses = []
                        for _, r in df.iterrows():
                            new_courses.append({
                                "id": str(r["id"]).strip(), "ad": str(r["ad"]).strip(),
                                "hoca": str(r["hoca"]).strip(), "sinif": int(r["sinif"]),
                                "sure": int(r["sure"]), "ardisik": bool(r["ardisik"]), "online": bool(r["online"])
                            })
                        if replace_all:
                            st.session_state.courses = new_courses
                        else:
                            by_id = {c["id"]: c for c in st.session_state.courses}
                            for nc in new_courses:
                                if nc["id"] in by_id and update_existing:
                                    by_id[nc["id"]].update(nc)
                                elif nc["id"] not in by_id:
                                    st.session_state.courses.append(nc)
                        st.success(f"İçe aktarma tamam: {len(new_courses)} ders okundu.")
                        st.rerun()
                except Exception as e:
                    st.error(f"İçe aktarma hatası: {e}")

    with st.expander("Takvim, Sınıflar ve Gün Penceresi", expanded=True):
        days_str = st.text_input("Günler (virgülle)", value=",".join(st.session_state.days))
        spd = st.number_input("Günlük slot sayısı", min_value=1, max_value=16,
                              value=st.session_state.slots_per_day, step=1)
        if st.button("Takvim Güncelle"):
            st.session_state.days = [d.strip() for d in days_str.split(",") if d.strip()]
            st.session_state.slots_per_day = int(spd)
            dcount = len(st.session_state.days)
            st.session_state.day_start_slot = {i: st.session_state.day_start_slot.get(i, 0) for i in range(dcount)}
            st.session_state.day_use_slots  = {i: st.session_state.day_use_slots.get(i, st.session_state.slots_per_day) for i in range(dcount)}
            st.session_state.time_labels = {i: st.session_state.time_labels.get(i, f"{9+i:02d}:00")
                                            for i in range(st.session_state.slots_per_day)}
            st.rerun()

        with st.form("slot_labels"):
            st.markdown("**Slot Saat Etiketleri**")
            tl = {}
            for i in range(st.session_state.slots_per_day):
                tl[i] = st.text_input(f"Slot {i+1}", value=st.session_state.time_labels.get(i, f"{9+i:02d}:00"))
            if st.form_submit_button("Etiketleri Kaydet"):
                st.session_state.time_labels = tl
                st.success("Kaydedildi.")

        st.markdown("**Sınıflar (Oda)**")
        rm_to_del = st.selectbox("Silmek için seç", options=["(seçme)"] + [r["id"] for r in st.session_state.rooms])
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Seçili sınıfı sil") and rm_to_del != "(seçme)":
                st.session_state.rooms = [r for r in st.session_state.rooms if r["id"] != rm_to_del]
                st.rerun()
        with c2:
            rid = st.text_input("Yeni sınıf ID")
            if st.button("Sınıf Ekle"):
                if rid and rid not in [r["id"] for r in st.session_state.rooms]:
                    st.session_state.rooms.append({"id": rid})
                    st.rerun()
        st.caption("Mevcut: " + ", ".join(r["id"] for r in st.session_state.rooms))

        st.markdown("---")
        st.markdown("### Gün Penceresi (Başlangıç Slotu & Kullanılacak Slot Sayısı)")
        dcount = len(st.session_state.days)
        for i in range(dcount):
            col1, col2, col3 = st.columns([0.35, 0.35, 0.3])
            with col1:
                st.write(f"**{st.session_state.days[i]}**")
            with col2:
                st.session_state.day_start_slot[i] = st.number_input(
                    f"{st.session_state.days[i]} başlangıç slotu", min_value=0,
                    max_value=st.session_state.slots_per_day-1, value=int(st.session_state.day_start_slot.get(i,0)),
                    key=f"start_{i}"
                )
            with col3:
                st.session_state.day_use_slots[i] = st.number_input(
                    f"{st.session_state.days[i]} kullanılacak slot sayısı", min_value=0,
                    max_value=st.session_state.slots_per_day,
                    value=int(st.session_state.day_use_slots.get(i, st.session_state.slots_per_day)),
                    key=f"use_{i}"
                )
        st.caption("Not: Gün penceresi dışında kalan slotlara ders yerleştirilmez.")

    with st.expander("Hocalar ve Uygunsuz Saatler", expanded=False):
        colh1, colh2 = st.columns(2)
        with colh1:
            new_inst = st.text_input("Yeni hoca adı")
            if st.button("Hoca Ekle"):
                if new_inst and new_inst not in st.session_state.instructors:
                    st.session_state.instructors.append(new_inst)
                    st.session_state.instructor_unavailable[new_inst] = set()
                    st.rerun()
        with colh2:
            del_inst = st.selectbox("Silinecek hoca", options=["(seçme)"] + st.session_state.instructors)
            if st.button("Hoca Sil") and del_inst != "(seçme)":
                st.session_state.instructors.remove(del_inst)
                st.session_state.instructor_unavailable.pop(del_inst, None)
                st.rerun()

        if st.session_state.instructors:
            sel = st.selectbox("Hoca seç", options=st.session_state.instructors)
            st.caption("Uygun DEĞİL slotları işaretleyin")
            days = st.session_state.days; spd = st.session_state.slots_per_day
            for d in range(len(days)):
                st.markdown(f"**{days[d]}**")
                cols = st.columns(spd)
                for s in range(spd):
                    key = f"inst_{sel}_{d}_{s}"
                    chk = (d, s) in st.session_state.instructor_unavailable.get(sel, set())
                    cols[s].checkbox(st.session_state.time_labels.get(s, f"{s+1}"), value=chk, key=key)
            if st.button("Kaydet (Hoca uygunluk)"):
                updated = set()
                for d in range(len(days)):
                    for s in range(spd):
                        if st.session_state.get(f"inst_{sel}_{d}_{s}"):
                            updated.add((d, s))
                st.session_state.instructor_unavailable[sel] = updated
                st.success("Güncellendi.")

    with st.expander("Dersler", expanded=True):
        st.table(st.session_state.courses)
        st.markdown("---")
        ids = ["(yeni)"] + [c["id"] for c in st.session_state.courses]
        choose = st.selectbox("Ders seç (düzenle)", options=ids)
        editing = next((c for c in st.session_state.courses if c["id"] == choose), None)
        if not editing:
            editing = {"id":"", "ad":"", "hoca": (st.session_state.instructors[0] if st.session_state.instructors else ""),
                       "sinif":1,"sure":1,"ardisik":False,"online":False}
        c1, c2 = st.columns(2)
        with c1:
            cid = st.text_input("Ders ID", value=editing["id"])
            cad = st.text_input("Ders Adı", value=editing["ad"])
            choca = st.selectbox("Hoca", options=st.session_state.instructors or [""],
                                 index=(st.session_state.instructors.index(editing["hoca"]) if editing["hoca"] in st.session_state.instructors else 0))
        with c2:
            csinif = st.number_input("Sınıf (1-4)", min_value=1, max_value=4, value=int(editing["sinif"]), step=1)
            csure  = st.number_input("Süre (slot)", min_value=1, max_value=10, value=int(editing["sure"]), step=1)
        card = st.toggle("Ardışık mı? (sure>1 ise)", value=bool(editing["ardisik"]))
        conline = st.toggle("Online mı?", value=bool(editing["online"]))
        b1, b2, b3 = st.columns(3)
        with b1:
            if st.button("Kaydet/Güncelle"):
                if not cid: st.error("ID boş olamaz.")
                else:
                    if choose != "(yeni)":
                        editing.update({"id":cid,"ad":cad,"hoca":choca,"sinif":int(csinif),
                                        "sure":int(csure),"ardisik":bool(card),"online":bool(conline)})
                    else:
                        st.session_state.courses.append({"id":cid,"ad":cad,"hoca":choca,"sinif":int(csinif),
                                                         "sure":int(csure),"ardisik":bool(card),"online":bool(conline)})
                st.rerun()
        with b2:
            if choose != "(yeni)" and st.button("Seçileni Sil"):
                st.session_state.courses = [c for c in st.session_state.courses if c["id"] != choose]
                st.rerun()
        with b3:
            if st.button("Tüm Listeyi Temizle"):
                st.session_state.courses = []
                st.rerun()

    with st.expander("📌 Pinler (Belirli slotlara sabitle)", expanded=True):
        if not st.session_state.courses:
            st.info("Önce ders ekleyin.")
        else:
            colp1, colp2 = st.columns(2)
            with colp1:
                pin_course = st.selectbox("Ders (ID)", options=[c["id"] for c in st.session_state.courses])
                pin_day = st.selectbox("Gün", options=list(range(len(st.session_state.days))),
                                       format_func=lambda i: st.session_state.days[i])
                pin_start = st.number_input("Başlangıç slotu", min_value=0,
                                            max_value=st.session_state.slots_per_day-1, value=0, step=1)
            with colp2:
                pin_channel = st.selectbox("Kanal", options=["FaceToFace","Online"])
                pin_room = ""
                if pin_channel == "FaceToFace":
                    pin_room = st.selectbox("Oda", options=[r["id"] for r in st.session_state.rooms])
                if st.button("Pin Ekle"):
                    new_pin = {"id": pin_course, "day": int(pin_day), "start": int(pin_start), "channel": pin_channel}
                    if pin_channel == "FaceToFace":
                        new_pin["room"] = pin_room
                    st.session_state.pins.append(new_pin)
                    st.success("Pin eklendi.")
                    st.rerun()

            st.markdown("**Mevcut Pinler**")
            if not st.session_state.pins:
                st.caption("Henüz pin yok.")
            else:
                pin_df = pd.DataFrame(st.session_state.pins)
                st.table(pin_df)
                del_idx = st.number_input("Silinecek pin indexi (0-based)", min_value=0,
                                          max_value=max(0, len(st.session_state.pins)-1),
                                          value=0, step=1)
                if st.button("Seçili pini sil"):
                    if st.session_state.pins:
                        st.session_state.pins.pop(int(del_idx))
                        st.success("Pin silindi.")
                        st.rerun()
                if st.button("Tüm pinleri temizle"):
                    st.session_state.pins = []
                    st.success("Tüm pinler temizlendi.")
                    st.rerun()

with right:
    st.header("Gün Gün Planlama")

    with st.expander("⚙️ Kısıt Ayarları ve Strateji", expanded=True):
        cs = st.session_state.constraint_settings
        col1, col2 = st.columns(2)
        with col1:
            online_cap = st.number_input("Slot başına MAKS. ONLINE", min_value=0, max_value=50,
                                         value=int(cs["online_cap"]), step=1)
            max_per_room = st.number_input("Slot/ODA başına MAKS. yüz yüze", min_value=1, max_value=5,
                                           value=int(cs["max_per_room"]), step=1)
        with col2:
            enf_inst = st.checkbox("Hoca aynı anda tek derste olsun", value=bool(cs["enf_instructor_no_overlap"]))
            enf_class = st.checkbox("Sınıf (1–4) aynı anda tek derste olsun", value=bool(cs["enf_class_no_overlap"]))
        st.session_state.strategy = st.selectbox(
            "Sıralama stratejisi",
            ["Kıtlık-önce (önerilir)", "Klasik: uzunluk-önce"],
            index=0 if st.session_state.strategy.startswith("Kıtlık") else 1
        )
        if st.button("Kısıtları Kaydet"):
            st.session_state.constraint_settings = {
                "online_cap": int(online_cap),
                "max_per_room": int(max_per_room),
                "enf_instructor_no_overlap": bool(enf_inst),
                "enf_class_no_overlap": bool(enf_class),
            }
            st.success("Kaydedildi.")

    if st.button("📅 GÜN GÜN PLANLA"):
        days = st.session_state.days
        spd = st.session_state.slots_per_day
        rooms = st.session_state.rooms
        courses = st.session_state.courses
        inst_unav = st.session_state.instructor_unavailable
        cs = st.session_state.constraint_settings

        timetable_df, diag_df, placed, unplaced = greedy_schedule(
            days, spd, rooms, courses, inst_unav, cs,
            day_start_slot=st.session_state.day_start_slot,
            day_use_slots=st.session_state.day_use_slots,
            pins=st.session_state.pins,
            strategy=st.session_state.strategy
        )

        placed_courses = len(set(ci for (ci,_,_,_,_) in placed))
        st.success(f"Yerleşen ders: {placed_courses}/{len(courses)}")

        st.subheader("Haftalık Tablo (Gün Gün)")
        render_day_tables(
            timetable_df,
            days=st.session_state.days,
            rooms=st.session_state.rooms,
            time_labels=st.session_state.time_labels
        )

        # CSV indir
        out = io.StringIO()
        timetable_df.to_csv(out, index=False)
        st.download_button("Programı CSV indir", data=out.getvalue(),
                           file_name="timetable.csv", mime="text/csv")

        # Excel indir (gün sayfası bazlı)
        excel_bytes = timetable_to_excel_bytes(
            timetable_df,
            days=st.session_state.days,
            rooms=st.session_state.rooms,
            time_labels=st.session_state.time_labels
        )
        st.download_button("📊 Programı Excel indir", data=excel_bytes.getvalue(),
                           file_name="timetable.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # PDF indir (wrap + dinamik satır)
        pdf_path = "timetable.pdf"
        timetable_to_pdf(
            timetable_df,
            days=st.session_state.days,
            rooms=st.session_state.rooms,
            time_labels=st.session_state.time_labels,
            pdf_path=pdf_path
        )
        with open(pdf_path, "rb") as f:
            st.download_button("📄 Programı PDF indir", data=f.read(),
                               file_name="timetable.pdf", mime="application/pdf")

        # Yerleşemeyenler
        st.subheader("Yerleşemeyen Dersler")
        if diag_df.empty:
            st.info("Tüm dersler yerleşti. 🎉")
        else:
            st.dataframe(diag_df, use_container_width=True)
            out2 = io.StringIO()
            diag_df.to_csv(out2, index=False)
            st.download_button("Yerleşemeyenler (CSV)", data=out2.getvalue(),
                               file_name="unscheduled_diagnostics.csv", mime="text/csv")

    st.markdown("---")
    st.caption("PDF'de taşma olmaması için metinler sütun genişliğine göre sarılır; satır yükseklikleri dinamik ayarlanır. Excel'de her gün ayrı sayfada wrap açık gelir.")
